import os
from skimage import io
from skimage.util import img_as_float32
from tqdm import tqdm

from utils.metrics import comp_upto_shif_rot_algn_color_multiprocess
import openpyxl
from openpyxl import Workbook

dataset = "MSCOCO"
method_name_list = ['Ours']
for method_name in method_name_list:
    print(f"正在处理 {method_name} 方法...")
    restored_image_dir = rf"nonuniform/{dataset}/{method_name}/restored"
    gt_image_dir = rf"nonuniform/{dataset}/ground_truth"
    # 写入Excel文件
    excel_path = "MSCOCO_nonuniform.xlsx"
    sheet_name = method_name

    data = []  # 存储结果数据

    for filename in tqdm(os.listdir(restored_image_dir)):
        if '_gyro_' not in filename:
            continue  # 跳过不含_gyro_的文件

        # 生成对应的ground truth文件名
        new_filename = filename.split('_gyro_')[0] + '.png'

        # 构建文件路径
        restored_image_path = os.path.join(restored_image_dir, filename)
        gt_image_path = os.path.join(gt_image_dir, new_filename)

        # 检查ground truth文件是否存在
        if not os.path.exists(gt_image_path):
            print(f"警告: 未找到匹配的GT文件 {new_filename}，跳过 {filename}")
            continue

        # 读取并转换图像
        try:
            restored_image = img_as_float32(io.imread(restored_image_path)[:, :, :3])
            gt_image = img_as_float32(io.imread(gt_image_path)[:, :, :3])
        except Exception as e:
            print(f"读取图像时出错: {str(e)}")
            continue

        # 计算质量指标
        try:
            output = comp_upto_shif_rot_algn_color_multiprocess(
                restored_image, gt_image,
                cut=15, maxshift=14, maxangle=0.5,
                shift_inter=1, angle_inter=0.1,
                ssim_compute=True
            )
            psnr = output[0]
            ssim = output[1]
        except Exception as e:
            print(f"计算指标时出错: {str(e)}")
            continue

        # 提取文件名（不含扩展名）
        base_name = os.path.splitext(filename)[0]
        data.append([base_name, psnr, ssim])

    # 处理Excel工作簿
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:  # 删除默认sheet
            wb.remove(wb["Sheet"])

    # 删除已存在的同名sheet
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # 创建新sheet并写入数据
    ws = wb.create_sheet(sheet_name)
    ws.append(["Filename", "PSNR", "SSIM"])  # 列标题
    for row in data:
        ws.append(row)

    # 保存文件
    wb.save(excel_path)
    print(f"结果已成功写入 {excel_path} 的 [{sheet_name}] 工作表中。")
