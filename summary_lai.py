import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取Excel文件
excel_file = "lai_nonuniform.xlsx"
xls = pd.ExcelFile(excel_file)

# 获取所有sheet名称并排除summary sheet
sheet_names = xls.sheet_names
summary_sheet_name = 'summary'
if summary_sheet_name in sheet_names:
    sheet_names.remove(summary_sheet_name)

# 定义类别
categories = ['Manmade', 'Natural', 'People', 'Saturated', 'Text']

# 计算每个方法的PSNR和SSIM平均值
summary_data = []
for sheet_name in sheet_names:
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    # 检查必要的列是否存在
    if not all(col in df.columns for col in ['Filename', 'PSNR', 'SSIM']):
        print(f"Sheet {sheet_name} 不包含所需的列。")
        continue

    # 初始化存储PSNR和SSIM的列表
    category_psnr = {cat: [] for cat in categories}
    category_ssim = {cat: [] for cat in categories}
    all_psnr = []
    all_ssim = []

    # 根据文件名分类并收集数据
    for index, row in df.iterrows():
        filename = row['Filename']
        psnr = row['PSNR']
        ssim = row['SSIM']

        # 假设文件名以类别开头
        for cat in categories:
            if filename.startswith(cat.lower()):
                category_psnr[cat].append(psnr)
                category_ssim[cat].append(ssim)
                break

        all_psnr.append(psnr)
        all_ssim.append(ssim)

    # 计算每个类别的平均值
    avg_psnr = {cat: sum(category_psnr[cat]) / len(category_psnr[cat]) if category_psnr[cat] else None for cat in
                categories}
    avg_ssim = {cat: sum(category_ssim[cat]) / len(category_ssim[cat]) if category_ssim[cat] else None for cat in
                categories}

    # 计算所有行的平均值（Average）
    avg_psnr['Average'] = sum(all_psnr) / len(all_psnr) if all_psnr else None
    avg_ssim['Average'] = sum(all_ssim) / len(all_ssim) if all_ssim else None

    # 存储结果
    summary_data.append({
        'Method': sheet_name,
        'PSNR_Manmade': avg_psnr['Manmade'],
        'SSIM_Manmade': avg_ssim['Manmade'],
        'PSNR_Natural': avg_psnr['Natural'],
        'SSIM_Natural': avg_ssim['Natural'],
        'PSNR_People': avg_psnr['People'],
        'SSIM_People': avg_ssim['People'],
        'PSNR_Saturated': avg_psnr['Saturated'],
        'SSIM_Saturated': avg_ssim['Saturated'],
        'PSNR_Text': avg_psnr['Text'],
        'SSIM_Text': avg_ssim['Text'],
        'PSNR_Average': avg_psnr['Average'],
        'SSIM_Average': avg_ssim['Average']
    })

# 加载现有的Excel文件并创建summary sheet
wb = load_workbook(excel_file)
if summary_sheet_name in wb.sheetnames:
    del wb[summary_sheet_name]
summary_sheet = wb.create_sheet(summary_sheet_name, 0)

# 设置summary sheet的标题
summary_sheet.merge_cells('A1:A2')
summary_sheet['A1'] = 'Method'

# 设置第一行类别标题（合并单元格）
categories_with_avg = ['Manmade', 'Natural', 'People', 'Saturated', 'Text', 'Average']
col_idx = 2
for cat in categories_with_avg:
    summary_sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
    summary_sheet.cell(row=1, column=col_idx).value = cat
    col_idx += 2

# 设置第二行PSNR和SSIM标题
col_idx = 2
for _ in categories_with_avg:
    summary_sheet.cell(row=2, column=col_idx).value = 'PSNR'
    summary_sheet.cell(row=2, column=col_idx + 1).value = 'SSIM'
    col_idx += 2

# 将数据写入summary sheet
summary_df = pd.DataFrame(summary_data)
columns_order = ['Method']
for cat in categories_with_avg:
    columns_order.extend([f'PSNR_{cat}', f'SSIM_{cat}'])
summary_df = summary_df[columns_order]

# for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=False), 3):
#     for c_idx, value in enumerate(row, 1):
#         if c_idx == 1:  # 方法名称
#             summary_sheet.cell(row=r_idx, column=c_idx).value = value
#         else:
#             if c_idx % 2 == 0:  # PSNR，保留2位小数
#                 summary_sheet.cell(row=r_idx, column=c_idx).value = f"{value:.2f}" if value is not None else None
#             else:  # SSIM，保留3位小数
#                 summary_sheet.cell(row=r_idx, column=c_idx).value = f"{value:.3f}" if value is not None else None
for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=False), 3):
    for c_idx, value in enumerate(row, 1):
        cell = summary_sheet.cell(row=r_idx, column=c_idx)
        if c_idx == 1:
            cell.value = value
        else:
            if isinstance(value, (int, float)):
                # 保持数值类型，设置数字格式
                cell.value = value
                if c_idx % 2 == 0:  # PSNR列
                    cell.number_format = '0.00'  # 两位小数
                else:  # SSIM列
                    cell.number_format = '0.000'  # 三位小数
            else:
                cell.value = value


# 设置单元格居中对齐
for row in summary_sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 保存文件
wb.save(excel_file)

print("Summary sheet 已成功生成！")
